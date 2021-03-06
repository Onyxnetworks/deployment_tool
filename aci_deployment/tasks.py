#  Base Functions
import openpyxl, json, yaml
from operator import itemgetter

# Custom Functions
from aci_deployment.scripts.ipg_search import *
from aci_deployment.scripts.endpoint_search import *
from aci_deployment.scripts.external_epg_deployment import *
from aci_deployment.scripts.contract_deployment import *
from aci_deployment.scripts.ipg_deployment import *
from aci_deployment.scripts.baseline import APIC_LOGIN
from index.scripts.baseline import get_base_url
# Celery Functions
from celery import shared_task


@shared_task
def ENDPOINT_SEARCH(base_urls, filter_default, username, password, search_string):
    RESULTS = []

    # Build URL List to search.
    url_list = []
    for url in base_urls['ACI']:
        url_list.append(base_urls['ACI'][url])

    ENDPOINT_LIST = GET_ENDPOINTS(url_list, filter_default, username, password)

    try:
        network = ipaddress.IPv4Network(search_string)
        request_type = 'subnet'
    except:
        request_type = 'epg_name'

    for i in ENDPOINT_LIST:
        if request_type == 'subnet':
            if IPNetwork(search_string) in IPNetwork(i['Subnet']) or IPNetwork(i['Subnet']) in IPNetwork(search_string):
                RESULTS.append({'Subnet': i['Subnet'], 'Locality': i['Locality'], 'Location': i['Location'],
                                'EPG': i['EPG'], 'Scope': i['Scope'], 'AppProfile':
                                    i['AppProfile'], 'Tenant': i['Tenant']})
        if request_type == 'epg_name':
            if search_string.upper() in i['EPG'].upper():
                RESULTS.append({'Subnet': i['Subnet'], 'Locality': i['Locality'], 'Location': i['Location'],
                                'EPG': i['EPG'], 'Scope': i['Scope'], 'AppProfile':
                                    i['AppProfile'], 'Tenant': i['Tenant']})

    # Sort the list by subnet before returning it.
    RESULTS = sorted(RESULTS, key=itemgetter('Subnet'), reverse=True)
    return RESULTS


@shared_task
def aci_ipg_search(base_urls, username, password, search_string):
    results = []

    # Build URL List to search.
    url_list = []
    for url in base_urls['ACI']:
        url_list.append(base_urls['ACI'][url])

    ipg_list = get_ipg(url_list, username, password)

    # List all EPG for selected
    # Convert  List into upper case
    ipg_list_upper = [element.upper() for element in ipg_list]
    search_string = search_string.upper()
    if [s for s in ipg_list_upper if search_string in s.upper()]:

        fvpathatt_list = get_fvpathatt(url_list, search_string, username, password)

        # loop through output, print and add EPG's to list
        i = 0
        for ipg_location in fvpathatt_list:
            location = ipg_location['location']
            for epg in ipg_location['response']['imdata']:
                results.append({'location': location, 'ipg':
                    epg['fvRsPathAtt']['attributes']['tDn'].split('/')[-1][8:].strip(']'), 'tenant':
                                    epg['fvRsPathAtt']['attributes']['dn'].split('/')[1][3:], 'app_prof':
                                    epg['fvRsPathAtt']['attributes']['dn'].split('/')[2][3:], 'epg':
                                    epg['fvRsPathAtt']['attributes']['dn'].split('/')[3][4:], 'encap':
                                    epg['fvRsPathAtt']['attributes']['encap']})
                i += 1

    return results


def ipg_deployment_open_yaml(file, location):
    data = yaml.safe_load(file)
    ipg_list = []
    index = 1
    for lines in data:

        # Required Fields
        line = index
        environment = lines['environment'].upper()
        node_1 = str(lines['node_1'])
        ports = lines['ports']
        speed = lines['speed']
        description = lines['description']
        mode = lines['mode'].upper()


        # Optional Fields
        if 'node_2' in lines:
            node_2 = str(lines['node_2'])
        else:
            node_2 = None

        if 'vpc' in lines:
            vpc = 'YES'
        else:
            vpc = 'NO'

        if 'port_channel_policy' in lines:
            port_channel_policy = lines['port_channel_policy'].upper()
        else:
            port_channel_policy = None


        if 'epg_list' in lines:
            epg_list = lines['epg_list']
        else:
            epg_list = {}


        if 'vmm' in lines:
            if lines['vmm'].upper() == 'YES':
                vmm = True
            else:
                vmm = False
        else:
            vmm = False

        ipg_list.append({'line': line, 'environment': environment, 'node_1': node_1, 'node_2': node_2,
                         'ports': ports, 'speed': speed, 'vpc': vpc, 'port_channel_policy': port_channel_policy,
                         'description': description, 'epg_list': epg_list, 'mode': mode, 'vmm': vmm})

        index += 1

    return ipg_list


def ipg_deployment_excel_open_workbook(file, location):
    wb = openpyxl.load_workbook(file, data_only=True)
    if location == 'UKDC1':
        py_ws = wb['ACI_DC1']
    elif location == 'UKDC2':
        py_ws = wb['ACI_DC2']
    elif location == 'LAB':
        py_ws = wb['ACI_LAB']

    index = 4
    ipg_list = []

    # Loops through the rows in the worksheet to build IPG information
    for row in py_ws.iter_rows(min_row=4, max_col=11):
        epg_list = []

        # Skip empty cells
        if row[1].value is None:
            continue

        environment = row[0].value
        node_1 = str(row[1].value)

        if row[2].value:
            node_2 = str(row[2].value)
        else:
            node_2 = None

        ports = row[3].value
        speed = row[4].value
        if row[5].value is None:
            vpc = 'NO'
        else:
            vpc = row[5].value.upper()

        port_channel_policy = row[6].value

        if not row[8].value is None:
            try:
                epg_list = json.loads(row[8].value)
            except:
                epg_list = {"Incorrect Format - Line-{0} make sure to use {1} ".format(index, '""'): ''}

        else:
            epg_list = {}

        mode = row[9].value.upper()

        if row[10].value.upper() == 'YES':
            vmm = True
        else:
            vmm = False

        description = row[7].value

        ipg_list.append({'line': index, 'environment': environment, 'node_1': node_1, 'node_2': node_2,
                         'ports': ports, 'speed': speed, 'vpc': vpc, 'port_channel_policy': port_channel_policy,
                         'description': description, 'epg_list': epg_list, 'mode': mode, 'vmm': vmm})

        index += 1

    return ipg_list


@shared_task
def ipg_deployment_validation(ipg_list, location, url_dict, username, password):
    headers = {'content-type': 'application/json'}
    output_log = []
    base_url = url_dict[location]
    error = False
    vpc_presant = False

    output_log.append({'Headers': "Validating Formatting in Workbook."})
    for ipg in ipg_list:

        if ipg['vmm'] and ipg['epg_list']:
            error = True
            output_log.append({'Errors': 'Line: {0} - cannot have VMM integrated IPG with static EPG mappings.'.format(str(ipg['line']))})

        if ipg['environment'] is None:
            error = True
            output_log.append({'Errors': 'Line: {0} - No environment defined.'.format(str(ipg['line']))})

        if ipg['vpc'] is None:
            error = True
            output_log.append({'Errors': 'Line: {0} - VPC Field not defined'.format(str(ipg['line']))})
            continue

        if ipg['vpc'] == 'YES':
            # Mark VPC Presant for later in script
            vpc_presant = True
            # If VPC = Yes check to make sure both nodes are defined.
            if ipg['node_2'] is None:
                error = True
                output_log.append({'Errors': 'Line: {0} - Type VPC selected but Node 2 not defined.'.format(str(ipg['line']))})
                continue

            # If VPC = Yes check to make sure switches are a VPC pair
            if int(ipg['node_2']) - int(ipg['node_1']) != 1:
                error = True
                output_log.append({'Errors': 'Line: {0} - Type VPC selected but switches are not a VPC Pair.'.format(str(ipg['line']))})

            # If VPC = Yes check that a port-channel policy is defined.
            if ipg['port_channel_policy'] is None:
                error = True
                output_log.append({'Errors': 'Line: {0} - Type VPC selected but no port-channel policy defined.'.format(str(ipg['line']))})
        if ipg['vpc'] == 'NO':
            # If VPC = No check to make sure node_2 has no value
            if not ipg['node_2'] is None:
                error = True
                output_log.append({'Errors': 'Line: {0} - Non VPC selected but two switches defined.'.format(str(ipg['line']))})
            # If VPC = No check to make sure port-channel policy has no value
            if not ipg['port_channel_policy'] is None:
                error = True
                output_log.append({'Errors': 'Line: {0} - Non VPC selected but port-channel policy defined.'.format(str(ipg['line']))})
        # Check to make sure speed is defined.
        if ipg['speed'] is None:
            error = True
            output_log.append({'Errors': 'Line: {0} - Interface speed not defined.'.format(str(ipg['line']))})

        # Check to make sure description is defined.
        if ipg['description'] is None:
            error = True
            output_log.append({'Errors': 'Line: {0} - Description not defined.'.format(str(ipg['line']))})

    if not error:
        output_log.append({'NotificationsSuccess': 'Workbook validated successfully'})

        # Go and check if switches exist on fabric.
        switch_list = []
        for switches in ipg_list:
            switch_list.append(switches['node_1'])
            if not switches['node_2'] is None:
                switch_list.append(switches['node_2'])

        excel_switch_list = list(set(switch_list))
        fabric_switch_list = []
        # Login to fabric
        output_log.append({'Headers': 'Connecting to APIC'})
        apic_cookie = APIC_LOGIN(base_url, username, password)
        if apic_cookie:
            output_log.append({'Notifications': 'Successfully generated authentication cookie'})
            output_log.append({'Headers': 'Checking if nodes presant in fabric.'})
            get_fabric_nodes_response = get_fabric_nodes(base_url, apic_cookie, headers, output_log)
            if get_fabric_nodes_response[0]:
                output_log = get_fabric_nodes_response[1]
                error = True

            else:
                for nodes in get_fabric_nodes_response[1]['imdata']:
                    fabric_switch_list.append(nodes['fabricNode']['attributes']['id'])

                for switch in excel_switch_list:
                    if switch not in fabric_switch_list:
                        error = True
                        output_log.append({'Errors': 'Node {0} not found in fabric.'.format(switch)})

            if not error:
                output_log.append({'NotificationsSuccess': 'Fabric switches validated successfully'})

            # Check if IPG's already exist.
            if not error:
                output_log.append({'Headers': "Checking if IPG's presant in fabric."})
                get_fabric_vpc_ipgs_response = get_fabric_vpc_ipgs(base_url, apic_cookie, headers, output_log)
                if get_fabric_vpc_ipgs_response[0]:
                    output_log = get_fabric_vpc_ipgs_response[1]
                    error = True

                if not error:
                    get_fabric_ipgs_response = get_fabric_ipgs(base_url, apic_cookie, headers, output_log)
                    if get_fabric_ipgs_response[0]:
                        output_log = get_fabric_ipgs_response[1]
                        error = True

                if not error:
                    accessIpg = True
                    for ipg in ipg_list:
                        # Set created flag to false
                        ipg['lsp_mapped'] = False
                        ipg['presant'] = False

                        # Build IPG Name
                        if location == 'UKDC1':
                            ipg_prefix = '1'
                        elif location == 'UKDC2':
                            ipg_prefix = '2'
                        elif location == 'LAB':
                            ipg_prefix = '2'
                        environment = ipg['environment']
                        node_1 = ipg['node_1']

                        port = ipg['ports'].split('/')[-1]
                        fromPort = port
                        toPort = port
                        # Format the port
                        if len(port) == 1:
                            port = '0' + port

                        if '-' in port:
                            # format for multiple Interfaces
                            port_start = port.split('-')[0]
                            port_end = port.split('-')[1]
                            if len(port_start) > 1:
                                fromPort = port_start
                            if len(port_end) > 1:
                                toPort = port_end
                            if len(port_start) == 1:
                                fromPort = port_start
                                port_start = '0' + port_start
                            if len(port_end) == 1:
                                toPort = port_end
                                port_end = '0' + port_end

                            port = port_start + '-' + port_end

                        port_settings = {}
                        port_settings['toPort'] = toPort
                        port_settings['fromPort'] = fromPort
                        port_settings['blockDescription'] = ipg['description']
                        port_settings['lspDescription'] = 'P{0}-{1}'.format(port, ipg['description'])
                        port_settings['block'] = ipg['ports']

                        if ipg['vpc'] == 'YES':

                            # Build IPG Name
                            node_2 = ipg['node_2']

                            ipg_name = '{0}{1}-VPC-{2}-{3}-P{4}_IPG'.format(environment, ipg_prefix, node_1, node_2,
                                                                            port)

                            port_settings['ipg'] = ipg_name
                            port_settings['ipgPrefix'] = 'accbundle'
                            port_settings['lsp'] = 'VPC-{0}-{1}_LSP'.format(node_1, node_2)

                            ipg_settings = {}

                            if ipg['vmm']:
                                ipg_settings['aep'] = 'uni/infra/attentp-{0}{1}-HYPERVISOR_AEP'.format(ipg['environment'], ipg_prefix)
                            else:
                                ipg_settings['aep'] = 'uni/infra/attentp-{0}{1}-ACCESS_AEP'.format(
                                    ipg['environment'], ipg_prefix)

                            ipg_settings['name'] = ipg_name
                            ipg_settings['speed'] = '{0}-AUTO_POL'.format(ipg['speed'])
                            ipg_settings['portChannelPolicy'] = ipg['port_channel_policy']

                            # Add port & ipg settings to item.
                            ipg['portSettings'] = port_settings
                            ipg['ipgSettings'] = ipg_settings

                            fabric_ipg_list = []
                            for fabric_ipg_name in get_fabric_vpc_ipgs_response[1]['imdata']:
                                fabric_ipg_list.append(fabric_ipg_name['infraAccBndlGrp']['attributes']['name'])

                            if ipg_name in fabric_ipg_list:
                                ipg['presant'] = True
                                output_log.append({'NotificationsWarning': "Line: {0} {1} already exists on the fabric and won't be created".format(str(ipg['line']), ipg_name)})
                            else:
                                output_log.append({'Notifications': "Line: {0} {1} will be created".format(str(ipg['line']), ipg_name)})
                        if ipg['vpc'] == 'NO':

                            # Set IPG to access (default used for all non VPC ports).
                            ipg_name = '{0}{1}-SVR-ACCESS-{2}_IPG'.format(environment, ipg_prefix, ipg['speed'] )
                            port_settings['ipgPrefix'] = 'accportgrp'
                            port_settings['ipg'] = ipg_name
                            port_settings['lsp'] = 'LFS{0}_LSP'.format(node_1)

                            ipg_settings = {}
                            ipg_settings['name'] = ipg_name

                            # Add port & ipg settings to item.
                            ipg['portSettings'] = port_settings
                            ipg['ipgSettings'] = ipg_settings

                            fabric_ipg_list = []
                            for fabric_ipg_name in get_fabric_ipgs_response[1]['imdata']:
                                fabric_ipg_list.append(fabric_ipg_name['infraAccPortGrp']['attributes']['name'])

                            if ipg_name in fabric_ipg_list:
                                ipg['presant'] = True
                            else:
                                error = True
                                accessIpg = False
                    if not accessIpg:
                        output_log.append({'Errors': 'Unable to locate Access IPG on fabric. IPG Name: {0}'.format(ipg_name)})

            # Checking if VPC domain are correct.
            if not error:
                output_log.append({'Headers': 'Checking VPC Pairs.'})
                get_vpc_domain_response = get_vpc_domain(base_url, apic_cookie, headers, output_log)

                if get_vpc_domain_response[0]:
                    output_log = get_vpc_domain_response[1]
                    error = True

                else:
                    vpc_list = []
                    for vpc in get_vpc_domain_response[1]['imdata']:
                        vpc_nodes = []
                        vpc_name = vpc['fabricExplicitGEp']['attributes']['name']

                        # Get VPC Domain Details
                        get_vpc_detail_response = get_vpc_domain_detail(base_url, apic_cookie, headers, output_log,
                                                                        vpc_name)
                        if get_vpc_detail_response[0]:
                            output_log = get_vpc_detail_response[1]
                            error = True

                        else:

                            for node_detail in get_vpc_detail_response[1]['imdata']:
                                if 'fabricNodePEp' in node_detail:
                                    node_id = int(node_detail['fabricNodePEp']['attributes']['id'])
                                    vpc_nodes.append(node_id)

                            # Sort list so lowest node ID is 1st.
                            vpc_nodes = sorted(vpc_nodes)
                            vpc_list.append(vpc_nodes)
                    for ipg in ipg_list:
                        if ipg['vpc'] == 'YES':
                            form_vpc = [int(ipg['node_1']), int(ipg['node_2'])]
                            if sorted(form_vpc) not in vpc_list:
                                error = True
                                output_log.append({'Errors': 'Line: {0} Node: {1} & {2} not configured for VPC.'.format(str(ipg['line']), ipg['node_1'], ipg['node_2'])})
            # Check if ports are already in use.
            if not error:
                output_log.append({'NotificationsSuccess': "VPC's validated successfully"})
                output_log.append({'Headers': 'Checking interface availability.'})

                for ipg in ipg_list:
                    node1 = ipg['node_1']
                    if ipg['vpc'] == 'YES':
                        node2 = ipg['node_2']

                        # Format the port
                        port = ipg['ports'].split('/')[-1]

                        if '-' in port:
                            # format for multiple Interfaces
                            start_card = ipg['ports'].split('/')[0]
                            end_card = ipg['ports'].split('/')[0]
                            start_port = port.split('-')[0]
                            end_port = port.split('-')[1]

                        else:
                            start_card = ipg['ports'].split('/')[0]
                            end_card = ipg['ports'].split('/')[0]
                            start_port = port
                            end_port = port

                        # Get port details from VPC LSP
                        vpc = True
                        get_lsp_detail_response = get_lsp_detail(base_url, apic_cookie, headers, output_log, vpc, node1, node2)

                        if get_lsp_detail_response[0]:
                            output_log = get_lsp_detail_response[1]
                            error = True

                        else:
                            for lsp_detail in get_lsp_detail_response[1]['imdata']:
                                if 'infraHPortS' in lsp_detail:
                                    # Get Port details
                                    dn = lsp_detail['infraHPortS']['attributes']['dn']

                                    get_port_detail_response = get_port_detail(base_url, apic_cookie, headers,
                                                                               output_log, dn)

                                    if get_port_detail_response[0]:
                                        output_log = get_port_detail_response[1]
                                        error = True

                                    else:
                                        for port_detail in get_port_detail_response[1]['imdata']:

                                            if 'infraPortBlk' in port_detail:
                                                fromCard = port_detail['infraPortBlk']['attributes']['fromCard']
                                                fromPort = port_detail['infraPortBlk']['attributes']['fromPort']
                                                toCard = port_detail['infraPortBlk']['attributes']['toCard']
                                                toPort = port_detail['infraPortBlk']['attributes']['toPort']

                                            if 'infraRsAccBaseGrp' in port_detail:
                                                excel_port = ipg['ports'].split('/')[-1]
                                                # Format the port
                                                if len(excel_port) == 1:
                                                    excel_port = '0' + excel_port
                                                if '-' in excel_port:
                                                    # format for multiple Interfaces
                                                    port_start = excel_port.split('-')[0]
                                                    port_end = excel_port.split('-')[1]
                                                    if len(port_start) == 1:
                                                        port_start = '0' + port_start
                                                    if len(port_end) == 1:
                                                        port_end = '0' + port_end
                                                    excel_port = port_start + '-' + port_end

                                                ipg_name = ipg['ipgSettings']['name']

                                                fabric_ipg_name = \
                                                port_detail['infraRsAccBaseGrp']['attributes']['tDn'].split(
                                                    'accbundle-')[-1]
                                                if ipg_name != fabric_ipg_name:
                                                    if start_card == fromCard and start_port == fromPort and end_card == toCard and end_port == toPort:
                                                        error = True
                                                        output_log.append({'Errors': 'Line: ' + str(
                                                            ipg[
                                                                'line']) + ' IPG Name missmatch between the Workbook and Fabric. ' + ipg_name + ' & ' + fabric_ipg_name})

                                                if ipg_name == fabric_ipg_name:
                                                    if start_card == fromCard and start_port == fromPort and end_card == toCard and end_port == toPort:
                                                        ipg['presant'] = True
                                                        ipg['lsp_mapped'] = True
                                                        output_log.append({'NotificationsWarning': 'Line: ' + str(
                                                            ipg[
                                                                'line']) + ' ports already provisioned, no IPG will be configured but any additional EPGs will be pusshed'})

                        # Get port details from node_1 LSP
                        vpc = False
                        nonvpc_node_1 = ipg['node_1']
                        nonvpc_node_2 = ''
                        get_lsp_detail_response_node1 = get_lsp_detail(base_url, apic_cookie, headers, output_log, vpc,
                                                                       nonvpc_node_1, nonvpc_node_2)

                        if get_lsp_detail_response_node1[0]:
                            output_log = get_lsp_detail_response_node1[1]
                            error = True

                        else:
                            for lsp_detail in get_lsp_detail_response_node1[1]['imdata']:
                                if 'infraHPortS' in lsp_detail:
                                    # Get Port details
                                    dn = lsp_detail['infraHPortS']['attributes']['dn']

                                    get_port_detail_response = get_port_detail(base_url, apic_cookie, headers,
                                                                               output_log, dn)

                                    if get_port_detail_response[0]:
                                        output_log = get_port_detail_response[1]
                                        error = True

                                    else:
                                        for port_detail in get_port_detail_response[1]['imdata']:
                                            if 'infraPortBlk' in port_detail:
                                                fromCard = port_detail['infraPortBlk']['attributes']['fromCard']
                                                fromPort = port_detail['infraPortBlk']['attributes']['fromPort']
                                                toCard = port_detail['infraPortBlk']['attributes']['toCard']
                                                toPort = port_detail['infraPortBlk']['attributes']['toPort']

                                                if start_card == fromCard and start_port == fromPort and end_card == toCard and end_port == toPort:
                                                    error = True
                                                    output_log.append({'Errors': 'Line: ' + str(ipg[
                                                                                                    'line']) + ' ports already provisioned as non VPC port on LFS' + nonvpc_node_1})

                        nonvpc_node_1 = ipg['node_2']
                        get_lsp_detail_response_node2 = get_lsp_detail(base_url, apic_cookie, headers, output_log, vpc,
                                                                       nonvpc_node_1, nonvpc_node_2)

                        if get_lsp_detail_response_node2[0]:
                            output_log = get_lsp_detail_response_node2[1]
                            error = True

                        else:
                            for lsp_detail in get_lsp_detail_response_node2[1]['imdata']:
                                if 'infraHPortS' in lsp_detail:
                                    # Get Port details
                                    dn = lsp_detail['infraHPortS']['attributes']['dn']

                                    get_port_detail_response = get_port_detail(base_url, apic_cookie, headers,
                                                                               output_log, dn)

                                    if get_port_detail_response[0]:
                                        output_log = get_port_detail_response[1]
                                        error = True

                                    else:
                                        for port_detail in get_port_detail_response[1]['imdata']:
                                            if 'infraPortBlk' in port_detail:
                                                fromCard = port_detail['infraPortBlk']['attributes']['fromCard']
                                                fromPort = port_detail['infraPortBlk']['attributes']['fromPort']
                                                toCard = port_detail['infraPortBlk']['attributes']['toCard']
                                                toPort = port_detail['infraPortBlk']['attributes']['toPort']

                                                if start_card == fromCard and start_port == fromPort and end_card == toCard and end_port == toPort:
                                                    error = True
                                                    output_log.append({'Errors': 'Line: ' + str(
                                                        ipg[
                                                            'line']) + ' ports already provisioned as non VPC port on LFS' + nonvpc_node_1})

                    if ipg['vpc'] == 'NO':
                        # Format the port
                        port = ipg['ports'].split('/')[-1]

                        if '-' in port:
                            # format for multiple Interfaces
                            start_card = ipg['ports'].split('/')[0]
                            end_card = ipg['ports'].split('/')[0]
                            start_port = port.split('-')[0]
                            end_port = port.split('-')[1]

                        else:
                            start_card = ipg['ports'].split('/')[0]
                            end_card = ipg['ports'].split('/')[0]
                            start_port = port
                            end_port = port

                        # Get port details from node_1 LSP
                        vpc = False
                        nonvpc_node_1 = node1
                        nonvpc_node_2 = ''
                        get_lsp_detail_response = get_lsp_detail(base_url, apic_cookie, headers, output_log, vpc,
                                                                 nonvpc_node_1, nonvpc_node_2)

                        if get_lsp_detail_response[0]:
                            output_log = get_lsp_detail_response[1]
                            error = True

                        else:
                            for lsp_detail in get_lsp_detail_response[1]['imdata']:
                                if 'infraHPortS' in lsp_detail:
                                    # Get Port details
                                    dn = lsp_detail['infraHPortS']['attributes']['dn']

                                    get_port_detail_response = get_port_detail(base_url, apic_cookie, headers,
                                                                               output_log, dn)

                                    if get_port_detail_response[0]:
                                        output_log = get_port_detail_response[1]
                                        error = True

                                    else:
                                        for port_detail in get_port_detail_response[1]['imdata']:
                                            if 'infraPortBlk' in port_detail:
                                                fromCard = port_detail['infraPortBlk']['attributes']['fromCard']
                                                fromPort = port_detail['infraPortBlk']['attributes']['fromPort']
                                                toCard = port_detail['infraPortBlk']['attributes']['toCard']
                                                toPort = port_detail['infraPortBlk']['attributes']['toPort']

                                                if start_card == fromCard and start_port == fromPort and end_card == toCard and end_port == toPort:
                                                    ipg['presant'] = True
                                                    ipg['lsp_mapped'] = True
                                                    output_log.append({'NotificationsWarning': 'Line: {0} ports already provisioned, no IPG will be configured but any additional EPGs will be pusshed'.format(str(ipg['line']))})

            # Check if EPGs are created.
            if not error:
                output_log.append({'Headers': "Checking if EPG's exist."})
                get_all_epg_response = get_all_epgs(base_url, apic_cookie, headers, output_log)
                if get_all_epg_response[0]:
                    output_log = get_all_epg_response[1]
                    error = True

                else:
                    epg_list = get_all_epg_response[1]['imdata']
                    # build epg list
                    for ipg in ipg_list:
                        for key in ipg['epg_list']:
                            if key is None:
                                continue
                            else:
                                if key.upper() not in (s['fvAEPg']['attributes']['name'].upper() for s in epg_list):
                                    error = True
                                    output_log.append({'Errors': '{0} EPG not presant in fabric.'.format(key)})



            if not error:
                output_log.append({'NotificationsSuccess': "EPG's validated successfully"})

        else:
            output_log.append({'Errors': 'Failed to connect to apic.'})

    return output_log, ipg_list


@shared_task
def ipg_deployment_post(ipg_list, location, url_dict, username, password):
    error = False
    base_url = url_dict[location]
    output_log = []
    headers = {'content-type': 'application/json'}

    # --------------------------------------------------------------------------#
    # Begin Configuration
    # --------------------------------------------------------------------------#
    output_log.append({'Headers': 'Starting IPG provisioning.'})
    output_log.append({'Headers': 'Connecting to APIC'})
    apic_cookie = APIC_LOGIN(base_url, username, password)
    if apic_cookie:
        output_log.append({'Notifications': 'Successfully generated authentication cookie'})
        output_log.append({'Headers': "Creating IPG's"})
        for ipg in ipg_list:

            # Create new VPC IPGs
            if ipg['vpc'] == 'YES' and not ipg['presant']:
                output_log.append({'Headers2': "Creating IPG's for line: {0}".format(ipg['line'])})
                # Post call to create IPG's
                post_create_ipg_response = post_create_ipg(base_url, apic_cookie, headers, output_log, ipg['ipgSettings'])
                error = post_create_ipg_response[0]
                output_log = post_create_ipg_response[1]

        if not error:
            # Map IPG to LSP
            output_log.append({'Headers': "Mapping IPG's to Interface profiles"})
            for ipg in ipg_list:
                if not ipg['lsp_mapped']:
                    output_log.append({'Headers2': "Mapping LSP's for line: {0}".format(ipg['line'])})

                    if ipg['vpc'] == 'YES':
                        # Post call to create switch profile ports.
                        post_create_lsp_port_response = post_create_lsp_port(base_url, apic_cookie, headers, output_log,
                                                                             ipg['portSettings'])

                        error = post_create_lsp_port_response[0]
                        output_log = post_create_lsp_port_response[1]

                    if ipg['vpc'] == 'NO':
                        # Post call to create switch profile ports.
                        post_create_lsp_port_response = post_create_lsp_port(base_url, apic_cookie, headers, output_log,
                                                                             ipg['portSettings'])

                        error = post_create_lsp_port_response[0]
                        output_log = post_create_lsp_port_response[1]

        if not error:
            output_log.append({'Headers': "Pushing EPG static bindings"})
            for ipg in ipg_list:
                if ipg['epg_list']:
                    output_log.append({'Headers2': 'Pushing Bindings for line: {0}'.format(ipg['line'])})

                if ipg['vpc'] == 'YES':
                    # GET EPG DN
                    for key, value in ipg['epg_list'].items():
                        if key is None:
                            continue
                        else:
                            epg_dn = get_epg_detail(base_url, apic_cookie, headers, key, output_log)
                            if epg_dn[0]:
                                output_log = epg_dn[1]
                                error = True
                            else:
                                epg_dn = epg_dn[1]['imdata'][0]['fvAEPg']['attributes']['dn']

                                binding_settings = {}
                                if ipg['mode'] == 'TRUNK':
                                    binding_settings['mode'] = 'regular'
                                if ipg['mode'] == 'ACCESS':
                                    binding_settings['mode'] = 'native'
                                binding_settings['vpc'] = True
                                binding_settings['encap'] = str(value)
                                binding_settings['tDn'] = epg_dn
                                binding_settings['epg_name'] = key
                                binding_settings['ipg_name'] = ipg['ipgSettings']['name']
                                binding_settings['lsptDn'] = 'topology/pod-1/protpaths-{0}-{1}/pathep-[{2}]'.format\
                                    (ipg['node_1'], ipg['node_2'], ipg['ipgSettings']['name'])

                                create_binding_response = post_create_static_binding(base_url, apic_cookie,
                                                                                         headers, output_log,
                                                                                         binding_settings)
                                error = create_binding_response[0]
                                output_log = create_binding_response[1]

                if ipg['vpc'] == 'NO':
                    # GET EPG DN
                    for key, value in ipg['epg_list'].items():
                        if key is None:
                            continue
                        else:
                            epg_dn = get_epg_detail(base_url, apic_cookie, headers, key, output_log)
                            if epg_dn[0]:
                                output_log = epg_dn[1]
                                error = True
                            else:
                                epg_dn = epg_dn[1]['imdata'][0]['fvAEPg']['attributes']['dn']

                                binding_settings = {}
                                if ipg['mode'] == 'TRUNK':
                                    binding_settings['mode'] = 'regular'
                                if ipg['mode'] == 'ACCESS':
                                    binding_settings['mode'] = 'native'
                                binding_settings['vpc'] = True
                                binding_settings['encap'] = str(value)
                                binding_settings['tDn'] = epg_dn
                                binding_settings['epg_name'] = key
                                binding_settings['ipg_name'] = ipg['ipgSettings']['name']
                                binding_settings['lsptDn'] = 'topology/pod-1/paths-{0}/pathep-[eth{1}]'.format\
                                    (ipg['node_1'], ipg['ports'])

                                create_binding_response = post_create_static_binding(base_url, apic_cookie,
                                                                                         headers, output_log,
                                                                                         binding_settings)
                                error = create_binding_response[0]
                                output_log = create_binding_response[1]



    else:
        output_log.append({'Errors': 'Failed to connect to apic.'})

    return output_log


def CONTRACT_DEPLOYMENT_EXCEL_OPEN_WORKBOOK(WORKBOOK, LOCATION):
    WB = openpyxl.load_workbook(WORKBOOK, data_only=True)
    if LOCATION == 'UKDC1':
        PY_WS = WB['ACI_DC1']
    elif LOCATION == 'UKDC2':
        PY_WS = WB['ACI_DC2']
    elif LOCATION == 'LAB':
        PY_WS = WB['ACI_LAB']

    SERVICE_LIST = []
    RULE_LIST = []
    INDEX = 1

    # Loops through the rows in the worksheet to build contract information
    for row in PY_WS.iter_rows(min_row=2, max_col=10):
        CONSUMER_IP_LIST = []
        PROVIDER_IP_LIST = []
        CONTRACT_NAME = row[2].value.upper()
        if row[8].value:
            CONSUMER_EPG = row[8].value.upper()
        else:
            CONSUMER_EPG = 'BLANK'
        if row[5].value:
            PROVIDER_EPG = row[5].value.upper()
        else:
            PROVIDER_EPG = 'BLANK'
        if row[9].value:
            CONSUMER_IP_LIST = row[9].value.split()
        else:
            pass
        if row[6].value:
            PROVIDER_IP_LIST = row[6].value.split()
        else:
            pass
        if row[3].value:
            SERVICE_LIST = row[3].value.upper().split()
        if row[7].value:
            CONSUMER_L3OUT = row[7].value.upper()
        else:
            CONSUMER_L3OUT = 'INTERNAL'
        if row[4].value:
            PROVIDER_L3OUT = row[4].value.upper()
        else:
            PROVIDER_L3OUT = 'INTERNAL'

        INDEX += 1
        RULE_LIST.append(
            {'LINE': INDEX, 'PROVIDER_L3OUT': PROVIDER_L3OUT, 'CONSUMER_L3OUT': CONSUMER_L3OUT, 'NAME': CONTRACT_NAME,
             'CONSUMER_EPG': CONSUMER_EPG, 'CONSUMER_IP': CONSUMER_IP_LIST, 'PROVIDER_EPG': PROVIDER_EPG,
             'PROVIDER_IP': PROVIDER_IP_LIST, 'SERVICE': SERVICE_LIST})
        SERVICE_LIST = []

    return RULE_LIST


def external_epg_open_yaml(file, location):
    data = yaml.safe_load(file)
    epg_list = []
    index = 1
    for lines in data:
        # Objects for combined EPG/Contract Form
        if 'contract' in lines:
            if 'provider_l3out' in lines:
                provider_l3out = lines['provider_l3out'].upper()
            else:
                provider_l3out = 'INTERNAL'

            if 'consumer_l3out' in lines:
                consumer_l3out = lines['consumer_l3out'].upper()
            else:
                consumer_l3out = 'INTERNAL'

            if 'consumer_epg' in lines:
                consumer_epg = lines['consumer_epg'].upper()
            else:
                consumer_epg = 'BLANK'

            if 'consumer_ip' in lines:
                consumer_ip = lines['consumer_ip']
                i = 0
                for ip in consumer_ip:
                    if len(str(ip).split('/')) <= 1:
                        subnet = '{0}/32'.format(str(ip))
                        consumer_ip[i] = subnet
                    i += 1
            else:
                consumer_ip = []

            if 'provider_epg' in lines:
                provider_epg = lines['provider_epg'].upper()
            else:
                provider_epg = 'BLANK'

            if 'provider_ip' in lines:
                provider_ip = lines['provider_ip']
                i = 0
                for ip in provider_ip:
                    if len(str(ip).split('/')) <= 1:
                        subnet = '{0}/32'.format(str(ip))
                        consumer_ip[i] = subnet
                    i += 1
            else:
                provider_ip = []

            epg_list.append({'LINE': index, 'PROVIDER_L3OUT': provider_l3out, 'CONSUMER_L3OUT': consumer_l3out,
                             'CONSUMER_EPG': consumer_epg, 'CONSUMER_IP': consumer_ip,
                             'PROVIDER_EPG': provider_epg, 'PROVIDER_IP': provider_ip})

            index += 1

        # Objects for combined EPG Only Form
        if 'l3out' in lines:
            # Only use Provicer settings to catch ant VIP requirements.
            consumer_ip = []
            consumer_l3out = 'INTERNAL'
            consumer_epg = 'BLANK'

            for items in lines['l3out']:
                for epgs in items['epg']:
                    provider_l3out = items['name'].upper()
                    provider_epg = epgs['name'].upper()

                    if 'subnets' in epgs:
                        provider_ip = epgs['subnets']
                        i = 0
                        for ip in provider_ip:
                            if len(str(ip).split('/')) <= 1:
                                subnet = '{0}/32'.format(str(ip))
                                provider_ip[i] = subnet
                            i += 1
                    else:
                        provider_ip = []

                    epg_list.append({'LINE': index, 'PROVIDER_L3OUT': provider_l3out, 'CONSUMER_L3OUT': consumer_l3out,
                                     'CONSUMER_EPG': consumer_epg, 'CONSUMER_IP': consumer_ip,
                                     'PROVIDER_EPG': provider_epg, 'PROVIDER_IP': provider_ip})

                    index += 1
    return epg_list


def EXTERNAL_EPG_EXCEL_OPEN_WORKBOOK(WORKBOOK, LOCATION):
    WB = openpyxl.load_workbook(WORKBOOK, data_only=True)
    if LOCATION == 'UKDC1':
        PY_WS = WB['ACI_DC1']
    elif LOCATION == 'UKDC2':
        PY_WS = WB['ACI_DC2']
    elif LOCATION == 'LAB':
        PY_WS = WB['ACI_LAB']

    RULE_LIST = []
    INDEX = 1

    # Loops through the rows in the worksheet to build contract information
    for row in PY_WS.iter_rows(min_row=2, max_col=10):
        CONSUMER_IP_LIST = []
        PROVIDER_IP_LIST = []
        if row[8].value:
            CONSUMER_EPG = row[8].value.upper()
        else:
            CONSUMER_EPG = 'BLANK'
        if row[5].value:
            PROVIDER_EPG = row[5].value.upper()
        else:
            PROVIDER_EPG = 'BLANK'
        if row[9].value:
            CONSUMER_IP_LIST = row[9].value.split()
            i = 0
            for IP in CONSUMER_IP_LIST:
                if len(IP.split('/')) <= 1:
                    IP_SUBNET = IP + '/32'
                    CONSUMER_IP_LIST[i] = IP_SUBNET
                i = i + 1
        else:
            pass
        if row[6].value:
            PROVIDER_IP_LIST = row[6].value.split()
            i = 0
            for IP in PROVIDER_IP_LIST:
                if len(IP.split('/')) <= 1:
                    IP_SUBNET = IP + '/32'
                    PROVIDER_IP_LIST[i] = IP_SUBNET
                i = i + 1
        else:
            pass
        if row[7].value:
            CONSUMER_L3OUT = row[7].value.upper()
        else:
            CONSUMER_L3OUT = 'INTERNAL'
        if row[4].value:
            PROVIDER_L3OUT = row[4].value.upper()
        else:
            PROVIDER_L3OUT = 'INTERNAL'
        INDEX += 1
        RULE_LIST.append({'LINE': INDEX, 'PROVIDER_L3OUT': PROVIDER_L3OUT, 'CONSUMER_L3OUT': CONSUMER_L3OUT,
                          'CONSUMER_EPG': CONSUMER_EPG, 'CONSUMER_IP': CONSUMER_IP_LIST,
                          'PROVIDER_EPG': PROVIDER_EPG, 'PROVIDER_IP': PROVIDER_IP_LIST})
    return RULE_LIST


@shared_task
def EXTERNAL_EPG_VALIDATION(RULE_LIST, location, url_dict, username, password):
    OUTPUT_LOG = []
    DISPLAY_LIST = []
    TENANT_LIST = ['RED', 'GREEN', 'BLUE']
    ERROR = False
    TENANT = 'common'

    BASE_URL = url_dict[location]

    OUTPUT_LOG.append({'Headers': 'Validating EPG names in Workbook.'})

    for rules in RULE_LIST:
        if rules['CONSUMER_EPG'] != 'BLANK' and rules['CONSUMER_L3OUT'] != 'INTERNAL':
            try:
                if len(rules['CONSUMER_EPG'].split('_')) > 2:
                    DISPLAY_LIST.append(rules['CONSUMER_EPG'])
                    ERROR = True

                elif rules['CONSUMER_EPG'].split('_')[1].upper() != 'EPG':
                    DISPLAY_LIST.append(rules['CONSUMER_EPG'])
                    ERROR = True

                elif rules['CONSUMER_EPG'].split('-')[0].upper() not in TENANT_LIST:
                    DISPLAY_LIST.append(rules['CONSUMER_EPG'])
                    ERROR = True

                elif rules['CONSUMER_L3OUT'].split('_')[0].endswith('CLOUD'):
                    continue

                # TEMP FIX FOR BLUE INET UNTIL RENAME
                elif rules['CONSUMER_L3OUT'] in ['BLUE-DC1-INET_L3O', 'BLUE-DC2-INET_L3O']:
                    continue
                elif not rules['CONSUMER_EPG'].split('_')[0].startswith(rules['CONSUMER_EPG'].split('_')[0]):
                    DISPLAY_LIST.append(rules['CONSUMER_EPG'])
                    ERROR = True

                else:
                    pass

            except:
                ERROR = True
                OUTPUT_LOG.append({'Errors': rules['CONSUMER_EPG'] + ' does not conform to the naming standard'})

        if rules['PROVIDER_EPG'] != 'BLANK' and rules['PROVIDER_L3OUT'] != 'INTERNAL':
            try:
                if len(rules['PROVIDER_EPG'].split('_')) > 2:
                    DISPLAY_LIST.append(rules['PROVIDER_EPG'])
                    ERROR = True

                elif rules['PROVIDER_EPG'].split('_')[1].upper() != 'EPG':
                    DISPLAY_LIST.append(rules['PROVIDER_EPG'])
                    ERROR = True

                elif rules['PROVIDER_EPG'].split('-')[0].upper() not in TENANT_LIST:
                    DISPLAY_LIST.append(rules['PROVIDER_EPG'])
                    ERROR = True

                elif rules['PROVIDER_L3OUT'].split('_')[0].endswith('CLOUD'):
                    continue

                # TEMP FIX FOR BLUE INET UNTIL RENAME
                elif rules['PROVIDER_L3OUT'] in ['BLUE-DC1-INET_L3O', 'BLUE-DC2-INET_L3O']:
                    continue
                elif not rules['PROVIDER_EPG'].split('_')[0].startswith(rules['PROVIDER_L3OUT'].split('_')[0]):
                    DISPLAY_LIST.append(rules['PROVIDER_EPG'])
                    ERROR = True

                else:
                    pass

            except:
                ERROR = True
                OUTPUT_LOG.append({'Errors': rules['PROVIDER_EPG'] + ' does not conform to the naming standard'})

        else:
            pass

    DISPLAY_SET = set(DISPLAY_LIST)
    for contracts in DISPLAY_SET:
        OUTPUT_LOG.append({'Errors': 'EPG "' + contracts + '" does not conform to the naming standard'})
    DISPLAY_LIST = []

    if not ERROR:
        OUTPUT_LOG.append({'NotificationsSuccess': 'EPG formatting validated successfully'})

    OUTPUT_LOG.append({'Headers': 'Validating IP addresses'})
    for addresses in RULE_LIST:
        if len(addresses['CONSUMER_IP']) >= 1:
            for subnets in addresses['CONSUMER_IP']:
                try:
                    network = ipaddress.IPv4Network(subnets)
                except ValueError:
                    ERROR = True
                    OUTPUT_LOG.append({'Errors': subnets + ' Is not a valid IPv4 address.'})

        if len(addresses['PROVIDER_IP']) >= 1:
            for subnets in addresses['PROVIDER_IP']:
                try:
                    network = ipaddress.IPv4Network(subnets)
                except ValueError:
                    ERROR = True
                    OUTPUT_LOG.append({'Errors': subnets + ' Is not a valid IPv4 address.'})

    if ERROR:
        OUTPUT_LOG.append({'Errors': 'Errors found in IP validation'})
    else:
        OUTPUT_LOG.append({'NotificationsSuccess': 'IP validation successful'})

        OUTPUT_LOG.append({'Headers': 'Connecting to APIC'})
        # Login to fabric
        APIC_COOKIE = APIC_LOGIN(BASE_URL, username, password)
        if APIC_COOKIE:
            OUTPUT_LOG.append({'Notifications': 'Successfully generated authentication cookie'})

            # Search for L3out and build URL to add IP's
            OUTPUT_LOG.append({'Headers': 'Validating L3Out Names'})
            L3OUT_LIST = []

            for rules in RULE_LIST:
                if rules['CONSUMER_L3OUT'] == 'INTERNAL':
                    pass
                else:
                    L3OUT_NAME = rules['CONSUMER_L3OUT']
                    L3OUT_SEARCH_RESPONSE = L3OUT_SEARCH(BASE_URL, APIC_COOKIE, TENANT, L3OUT_NAME, HEADERS)
                    if int(L3OUT_SEARCH_RESPONSE[0]['totalCount']) == 1:
                        if rules['CONSUMER_L3OUT'] == L3OUT_SEARCH_RESPONSE[0]['imdata'][0]['l3extOut']['attributes'][
                            'name']:
                            pass
                        else:
                            ERROR = True

                    else:
                        L3OUT_LIST.append(L3OUT_NAME)
                        ERROR = True

            for rules in RULE_LIST:
                if rules['PROVIDER_L3OUT'] == 'INTERNAL':
                    pass
                else:
                    L3OUT_NAME = rules['PROVIDER_L3OUT']
                    L3OUT_SEARCH_RESPONSE = L3OUT_SEARCH(BASE_URL, APIC_COOKIE, TENANT, L3OUT_NAME, HEADERS)
                    if int(L3OUT_SEARCH_RESPONSE[0]['totalCount']) == 1:
                        if rules['PROVIDER_L3OUT'] == L3OUT_SEARCH_RESPONSE[0]['imdata'][0]['l3extOut']['attributes'][
                            'name']:
                            pass
                        else:
                            ERROR = True

                    else:
                        L3OUT_LIST.append(L3OUT_NAME)
                        ERROR = True

            L3OUT_SET = set(L3OUT_LIST)
            for l3out in L3OUT_SET:
                OUTPUT_LOG.append({'Errors': 'L3Out: ' + l3out + ' Does not exist, please check naming.'})
            L3OUT_LIST = []

            if ERROR:
                OUTPUT_LOG.append({'Errors': 'Errors found in L3Out validation'})
            else:
                OUTPUT_LOG.append({'NotificationsSuccess': 'L3Out validation successful'})

                # Check if IP already exists in Same L3Out or same VRF
                OUTPUT_LOG.append({'Headers': 'Checking if IP currently exists within VRF'})

                # Get L3out VRF
                for rules in RULE_LIST:
                    OUTPUT_LOG.append({'Headers2': 'Checking subnets for line: ' + str(rules['LINE'])})
                    if rules['CONSUMER_L3OUT'] != 'INTERNAL' and rules['CONSUMER_EPG'] != 'BLANK':
                        L3OUT_NAME = rules['CONSUMER_L3OUT']
                        L3OUT_SEARCH_RESPONSE = L3OUT_SEARCH(BASE_URL, APIC_COOKIE, TENANT, L3OUT_NAME, HEADERS)
                        L3OUT_DATA = L3OUT_SEARCH_RESPONSE[1]['imdata']
                        L3OUT_SUBNETS = []
                        # Loop through the VRF pull out all other L3Outs and add any l3extsubnet to a list
                        for key in L3OUT_DATA:
                            # For Python 3+
                            if 'l3extRsEctx' in key:
                                # For Python 2.7
                                # if key.keys() == ['l3extRsEctx']:
                                VRF_DN = key['l3extRsEctx']['attributes']['tDn']
                                VRF_SEARCH_RESPONSE = VRF_SEARCH(BASE_URL, APIC_COOKIE, VRF_DN, HEADERS)
                                for vrf_l3o in VRF_SEARCH_RESPONSE['imdata']:
                                    # For Python 3+
                                    if 'fvRtEctx' in vrf_l3o:
                                        # For Python 2.7
                                        # if vrf_l3o.keys() == ['fvRtEctx']:
                                        L3OUT_DN = vrf_l3o['fvRtEctx']['attributes']['tDn']
                                        SUBNET_SEARCH_RESPONSE = SUBNET_SEARCH(BASE_URL, APIC_COOKIE, L3OUT_DN, HEADERS)
                                        for subnets in SUBNET_SEARCH_RESPONSE['imdata']:
                                            L3OUT_SUBNETS.append(subnets['l3extSubnet']['attributes']['ip'])
                                            EXISTING_SUBNET = subnets['l3extSubnet']['attributes']['ip']
                                            EXISTING_L3OUT = subnets['l3extSubnet']['attributes']['dn'].split('/')[2][
                                                             4:]
                                            EXISTING_EPG = subnets['l3extSubnet']['attributes']['dn'].split('/')[3][6:]
                                            SCOPE = subnets['l3extSubnet']['attributes']['scope'].split(',')
                                            if EXISTING_SUBNET in rules['CONSUMER_IP']:
                                                if 'import-security' in SCOPE:
                                                    if EXISTING_EPG == rules['CONSUMER_EPG']:
                                                        rules['CONSUMER_IP'].remove(EXISTING_SUBNET)
                                                    else:
                                                        OUTPUT_LOG.append({
                                                                              'Errors': EXISTING_SUBNET + ' already exists within ' + EXISTING_L3OUT + ' under EPG ' + EXISTING_EPG + ' no subnet configuration for this epg will be pushed.'})
                                                        ERROR = True
                                                        rules['CONSUMER_IP'].remove(EXISTING_SUBNET)

                                            else:
                                                for rule_subnet in rules['CONSUMER_IP']:
                                                    if IPNetwork(rule_subnet) in IPNetwork(EXISTING_SUBNET):
                                                        if int(EXISTING_SUBNET.split('/')[
                                                                   1]) >= 20 and 'import-security' in SCOPE:
                                                            OUTPUT_LOG.append({
                                                                                  'Errors': rule_subnet + ' already exists within ' + EXISTING_L3OUT + ' under EPG ' + EXISTING_EPG + ' inside ' + EXISTING_SUBNET})
                                                            ERROR = True
                                                            rules['CONSUMER_IP'].remove(rule_subnet)

                        if len(rules['CONSUMER_IP']) >= 1:
                            OUTPUT_LOG.append({'Notifications': 'The Following subnets will be added to the EPG: ' +
                                                                rules['CONSUMER_EPG']})
                            OUTPUT_LOG.append({'Notifications': str(rules['CONSUMER_IP'])})

                        else:
                            OUTPUT_LOG.append(
                                {'Notifications': 'No subnets will be added to EPG: ' + rules['CONSUMER_EPG']})

                    if rules['PROVIDER_L3OUT'] != 'INTERNAL' and rules['PROVIDER_EPG'] != 'BLANK':
                        L3OUT_NAME = rules['PROVIDER_L3OUT']
                        L3OUT_SEARCH_RESPONSE = L3OUT_SEARCH(BASE_URL, APIC_COOKIE, TENANT, L3OUT_NAME, HEADERS)
                        L3OUT_DATA = L3OUT_SEARCH_RESPONSE[1]['imdata']
                        L3OUT_SUBNETS = []
                        # Loop through the VRF pull out all other L3Outs and add any l3extsubnet to a list
                        for key in L3OUT_DATA:
                            # For Python 3+
                            if 'l3extRsEctx' in key:
                                # For Python 2.7
                                # if key.keys() == ['l3extRsEctx']:
                                VRF_DN = key['l3extRsEctx']['attributes']['tDn']
                                VRF_SEARCH_RESPONSE = VRF_SEARCH(BASE_URL, APIC_COOKIE, VRF_DN, HEADERS)
                                for vrf_l3o in VRF_SEARCH_RESPONSE['imdata']:
                                    # For Python 3+
                                    if 'fvRtEctx' in vrf_l3o:
                                        # For Python 2.7
                                        # if vrf_l3o.keys() == ['fvRtEctx']:
                                        L3OUT_DN = vrf_l3o['fvRtEctx']['attributes']['tDn']
                                        SUBNET_SEARCH_RESPONSE = SUBNET_SEARCH(BASE_URL, APIC_COOKIE, L3OUT_DN, HEADERS)
                                        for subnets in SUBNET_SEARCH_RESPONSE['imdata']:
                                            L3OUT_SUBNETS.append(subnets['l3extSubnet']['attributes']['ip'])
                                            EXISTING_SUBNET = subnets['l3extSubnet']['attributes']['ip']
                                            EXISTING_L3OUT = subnets['l3extSubnet']['attributes']['dn'].split('/')[2][
                                                             4:]
                                            EXISTING_EPG = subnets['l3extSubnet']['attributes']['dn'].split('/')[3][6:]
                                            SCOPE = subnets['l3extSubnet']['attributes']['scope'].split(',')
                                            if EXISTING_SUBNET in rules['PROVIDER_IP']:
                                                if 'import-security' in SCOPE:
                                                    if EXISTING_EPG == rules['PROVIDER_EPG']:
                                                        rules['PROVIDER_IP'].remove(EXISTING_SUBNET)

                                                    else:
                                                        OUTPUT_LOG.append({
                                                                              'Errors': EXISTING_SUBNET + ' already exists within ' + EXISTING_L3OUT + ' under EPG ' + EXISTING_EPG + ' no subnet configuration for this epg will be pushed.'})
                                                        ERROR = True
                                                        rules['PROVIDER_IP'].remove(EXISTING_SUBNET)

                                            else:
                                                for rule_subnet in rules['PROVIDER_IP']:
                                                    if IPNetwork(rule_subnet) in IPNetwork(EXISTING_SUBNET):
                                                        if int(EXISTING_SUBNET.split('/')[
                                                                   1]) >= 20 and 'import-security' in SCOPE:
                                                            OUTPUT_LOG.append({
                                                                                  'Errors': rule_subnet + ' already exists within ' + EXISTING_L3OUT + ' under EPG ' + EXISTING_EPG + ' inside ' + EXISTING_SUBNET})
                                                            ERROR = True
                                                            rules['PROVIDER_IP'].remove(rule_subnet)

                        if len(rules['PROVIDER_IP']) >= 1:
                            OUTPUT_LOG.append({'Notifications': 'The Following subnets will be added to the EPG: ' +
                                                                rules['PROVIDER_EPG']})
                            OUTPUT_LOG.append({'Notifications': str(rules['PROVIDER_IP'])})

                        else:
                            OUTPUT_LOG.append(
                                {'Notifications': 'No subnets will be added to EPG: ' + rules['PROVIDER_EPG']})

                # Search for VIPs
                OUTPUT_LOG.append({'Headers': 'Checking if any EPGs are for VIPS'})

                for rules in RULE_LIST:
                    if rules['PROVIDER_EPG'].split('_')[0].endswith('VS') and rules['PROVIDER_L3OUT'].endswith(
                            'DCI_L3O'):
                        for subnets in rules['PROVIDER_IP']:
                            if len(subnets.split('/')) != 0:
                                subnet = subnets.split('/')[0]
                            else:
                                subnet = subnets
                            if not ipaddress.ip_address(subnet).is_private:
                                OUTPUT_LOG.append(
                                    {'Notifications': rules['PROVIDER_EPG'] + ' contains a public address.'})
                                OUTPUT_LOG.append({
                                                      'Notifications': subnets + ' will be imported under the DCI and exported under the INET L3Outs'})


                            elif ipaddress.ip_address(subnet).is_private:
                                OUTPUT_LOG.append(
                                    {'Notifications': rules['PROVIDER_EPG'] + ' contains a private address.'})
                                OUTPUT_LOG.append({'Notifications': subnets + ' will be imported under the DCI L3Out'})

        else:
            OUTPUT_LOG.append({'Errors': 'Unable to connect to APIC.'})

    if not ERROR:
        OUTPUT_LOG.append({'ValidationSuccess': 'APIC Configuration validated successfully'})

    return OUTPUT_LOG


@shared_task
def EXTERNAL_EPG_DEPLOYMENT(RULE_LIST, location, url_dict, username, password):
    BASE_URL = url_dict[location]

    if location == 'UKDC1':
        DC = 'DC1'
    elif location == 'UKDC2':
        DC = 'DC2'
    elif location == 'LAB':
        DC = 'DC1'

    TENANT = 'common'
    OUTPUT_LOG = []
    HEADERS = {'content-type': 'application/json'}
    # --------------------------------------------------------------------------#
    # Begin Configuration
    # --------------------------------------------------------------------------#
    OUTPUT_LOG.append({'Headers': 'Starting External EPG Deployment.'})
    OUTPUT_LOG.append({'Headers': 'Connecting to APIC'})
    APIC_COOKIE = APIC_LOGIN(BASE_URL, username, password)
    if APIC_COOKIE:
        OUTPUT_LOG.append({'Notifications': 'Successfully generated authentication cookie'})

        check_i = 0
        for rules in RULE_LIST:
            check_i = check_i + 1
            L3OUT_CONSUME_EPG_CREATED = False
            L3OUT_PROVIDE_EPG_CREATED = False
            OUTPUT_LOG.append({'Headers2': 'Adding EPGs & Subnets for line: ' + str(rules['LINE'])})
            if rules['CONSUMER_L3OUT'] != 'INTERNAL' and rules['CONSUMER_EPG'] != 'BLANK':
                EPG_NAME = rules['CONSUMER_EPG']
                L3OUT_NAME = rules['CONSUMER_L3OUT']
                EXTERNAL_EPG_SEARCH_RESPONSE = EXTERNAL_EPG_SEARCH(BASE_URL, APIC_COOKIE, TENANT, L3OUT_NAME, EPG_NAME,
                                                                   HEADERS)
                if int(EXTERNAL_EPG_SEARCH_RESPONSE['totalCount']) == 1:
                    OUTPUT_LOG.append({
                                          'NotificationsWarning': 'EPG: ' + EPG_NAME + ' already exists under ' + L3OUT_NAME + ' and wont be created'})
                    L3OUT_CONSUME_EPG_CREATED = True
                if not L3OUT_CONSUME_EPG_CREATED:
                    OUTPUT_LOG.append(
                        {'Notifications': 'Adding External EPG: ' + EPG_NAME + ' TO L3Out: ' + L3OUT_NAME})
                    EXTERNAL_EPG_ADD(BASE_URL, APIC_COOKIE, TENANT, L3OUT_NAME, EPG_NAME, HEADERS, OUTPUT_LOG)
                # Add subnets to external EPG
                if len(rules['CONSUMER_IP']) != 0:
                    OUTPUT_LOG.append({'Notifications': 'Adding Subnets to External EPG: '})
                    for IP in rules['CONSUMER_IP']:
                        # SEARCH TO SEE IF SUBNET ALREADY EXISTS IN EPG:
                        L3OUT_DN = 'uni/tn-common/out-' + L3OUT_NAME + '/instP-' + EPG_NAME + '/extsubnet-[' + IP + ']'
                        SUBNET_SEARCH_RESPONSE = SUBNET_SEARCH_EXACT(BASE_URL, APIC_COOKIE, L3OUT_DN, HEADERS)
                        if int(SUBNET_SEARCH_RESPONSE['totalCount']) == 1:
                            for subnets in SUBNET_SEARCH_RESPONSE['imdata']:
                                EXISTING_SUBNET = subnets['l3extSubnet']['attributes']['ip']
                                SCOPE = subnets['l3extSubnet']['attributes']['scope'].split(',')
                                if EXISTING_SUBNET == IP:
                                    # IP Already Configured under EPG
                                    pass

                        else:
                            SCOPE = 'import-security'
                            EXTERNAL_EPG_SUBNET_ADD(BASE_URL, APIC_COOKIE, TENANT, L3OUT_NAME, EPG_NAME, HEADERS, IP,
                                                    SCOPE, OUTPUT_LOG)

            if rules['PROVIDER_L3OUT'] != 'INTERNAL' and rules['PROVIDER_EPG'] != 'BLANK':
                EPG_NAME = rules['PROVIDER_EPG']
                L3OUT_NAME = rules['PROVIDER_L3OUT']
                EXTERNAL_EPG_SEARCH_RESPONSE = EXTERNAL_EPG_SEARCH(BASE_URL, APIC_COOKIE, TENANT, L3OUT_NAME, EPG_NAME,
                                                                   HEADERS)
                if int(EXTERNAL_EPG_SEARCH_RESPONSE['totalCount']) == 1:
                    OUTPUT_LOG.append({
                                          'NotificationsWarning': 'EPG: ' + EPG_NAME + ' already exists under ' + L3OUT_NAME + ' and wont be created'})
                    L3OUT_PROVIDE_EPG_CREATED = True
                if not L3OUT_PROVIDE_EPG_CREATED:
                    OUTPUT_LOG.append(
                        {'Notifications': 'Adding External EPG: ' + EPG_NAME + ' TO L3Out: ' + L3OUT_NAME})
                    EXTERNAL_EPG_ADD(BASE_URL, APIC_COOKIE, TENANT, L3OUT_NAME, EPG_NAME, HEADERS, OUTPUT_LOG)
                # Add subnets to external EPG
                if len(rules['PROVIDER_IP']) != 0:
                    OUTPUT_LOG.append({'Notifications': 'Adding Subnets to External EPG: '})
                    for IP in rules['PROVIDER_IP']:
                        # SEARCH TO SEE IF SUBNET ALREADY EXISTS IN EPG:
                        L3OUT_DN = 'uni/tn-common/out-' + L3OUT_NAME + '/instP-' + EPG_NAME + '/extsubnet-[' + IP + ']'
                        SUBNET_SEARCH_RESPONSE = SUBNET_SEARCH_EXACT(BASE_URL, APIC_COOKIE, L3OUT_DN, HEADERS)
                        if int(SUBNET_SEARCH_RESPONSE['totalCount']) == 1:
                            for subnets in SUBNET_SEARCH_RESPONSE['imdata']:
                                EXISTING_SUBNET = subnets['l3extSubnet']['attributes']['ip']
                                SCOPE = subnets['l3extSubnet']['attributes']['scope'].split(',')
                                if EXISTING_SUBNET == IP:
                                    # IP Already Configured under EPG
                                    pass

                        else:
                            if len(IP.split('/')) != 0:
                                subnet = IP.split('/')[0]
                            else:
                                subnet = IP
                            # Check for VS EPGs
                            if rules['PROVIDER_EPG'].split('_')[0].endswith('VS') and rules['PROVIDER_L3OUT'].endswith(
                                    'DCI_L3O'):

                                if not ipaddress.ip_address(subnet).is_private:
                                    # Import Under DCI
                                    SCOPE = 'import-rtctrl,import-security'
                                    L3OUT_NAME = rules['PROVIDER_L3OUT']
                                    EPG_NAME = rules['PROVIDER_EPG']
                                    EXTERNAL_EPG_SUBNET_ADD(BASE_URL, APIC_COOKIE, TENANT, L3OUT_NAME, EPG_NAME,
                                                            HEADERS, IP, SCOPE, OUTPUT_LOG)

                                    # Export Under Inet
                                    SCOPE = 'export-rtctrl'
                                    # Build L3out name for Inet L3Out
                                    # Temp fix for BLUE INET
                                    if rules['PROVIDER_L3OUT'].split('-')[0] == 'BLUE':
                                        L3OUT_NAME = rules['PROVIDER_L3OUT'].split('-')[0] + '-' + DC + '-INET_L3O'
                                    else:
                                        L3OUT_NAME = rules['PROVIDER_L3OUT'].split('-')[0] + '-INET_L3O'

                                    EPG_NAME = L3OUT_NAME.split('_')[0] + '-ROUTING_EPG'
                                    EXTERNAL_EPG_SUBNET_ADD(BASE_URL, APIC_COOKIE, TENANT, L3OUT_NAME, EPG_NAME,
                                                            HEADERS, IP, SCOPE, OUTPUT_LOG)

                                if ipaddress.ip_address(subnet).is_private:
                                    # Import Under DCI
                                    L3OUT_NAME = rules['PROVIDER_L3OUT']
                                    EPG_NAME = rules['PROVIDER_EPG']
                                    SCOPE = 'import-rtctrl,import-security'
                                    EXTERNAL_EPG_SUBNET_ADD(BASE_URL, APIC_COOKIE, TENANT, L3OUT_NAME, EPG_NAME,
                                                            HEADERS, IP, SCOPE, OUTPUT_LOG)

                            else:
                                L3OUT_NAME = rules['PROVIDER_L3OUT']
                                EPG_NAME = rules['PROVIDER_EPG']
                                SCOPE = 'import-security'
                                EXTERNAL_EPG_SUBNET_ADD(BASE_URL, APIC_COOKIE, TENANT, L3OUT_NAME, EPG_NAME, HEADERS,
                                                        IP, SCOPE, OUTPUT_LOG)

    else:
        OUTPUT_LOG.append({'Errors': 'Unable to connect to APIC.'})

    return OUTPUT_LOG


@shared_task
def CONTRACT_DEPLOYMENT_VALIDATION(RULE_LIST, location, url_dict, username, password):
    CONTRACT_LIST = []
    FILTER_LIST = []
    DISPLAY_LIST = []
    OUTPUT_LOG = []
    ERROR = False
    HEADERS = {'content-type': 'application/json'}
    TENANT_LIST = ['RED', 'GREEN', 'BLUE']

    BASE_URL = url_dict[location]

    # Validate Contract Name formatting
    OUTPUT_LOG.append({'Headers': 'Validating Contract names in Workbook.'})

    try:
        for rules in RULE_LIST:
            if (len(rules['NAME'].split('_'))) > 2:
                DISPLAY_LIST.append(rules['NAME'])
                ERROR = True
            elif rules['NAME'].split('_')[1].upper() != 'GCTR':
                DISPLAY_LIST.append(rules['NAME'])
                ERROR = True
            elif rules['NAME'].split('-')[0].upper() not in TENANT_LIST:
                DISPLAY_LIST.append(rules['NAME'])
                ERROR = True

            else:
                pass

        DISPLAY_SET = set(DISPLAY_LIST)
        for contracts in DISPLAY_SET:
            OUTPUT_LOG.append({'Errors': 'Contract "' + contracts + '" does not conform to the naming standard'})
        DISPLAY_LIST = []

    except:
        OUTPUT_LOG.append({'Errors': 'Errors validating Contract names'})
        ERROR = True

    if not ERROR:
        OUTPUT_LOG.append({'NotificationsSuccess': 'Contract formatting validated successfully'})

    OUTPUT_LOG.append({'Headers': 'Validating EPG names in Workbook.'})

    for rules in RULE_LIST:
        if rules['CONSUMER_EPG'] != 'BLANK':
            try:
                if len(rules['CONSUMER_EPG'].split('_')) > 2:
                    DISPLAY_LIST.append(rules['CONSUMER_EPG'])
                    ERROR = True
                elif rules['CONSUMER_EPG'].split('_')[1].upper() != 'EPG':
                    DISPLAY_LIST.append(rules['CONSUMER_EPG'])
                    ERROR = True
                elif rules['CONSUMER_EPG'].split('-')[0].upper() not in TENANT_LIST:
                    DISPLAY_LIST.append(rules['CONSUMER_EPG'])
                    ERROR = True

                else:
                    pass

            except:
                ERROR = True
                OUTPUT_LOG.append(
                    {'Errors': 'EPG "' + rules['CONSUMER_EPG'] + '" does not conform to the naming standard'})

        if rules['PROVIDER_EPG'] != 'BLANK':
            try:
                if len(rules['PROVIDER_EPG'].split('_')) > 2:
                    DISPLAY_LIST.append(rules['PROVIDER_EPG'])
                    ERROR = True

                elif rules['PROVIDER_EPG'].split('_')[1].upper() != 'EPG':
                    DISPLAY_LIST.append(rules['PROVIDER_EPG'])
                    ERROR = True

                elif rules['PROVIDER_EPG'].split('-')[0].upper() not in TENANT_LIST:
                    DISPLAY_LIST.append(rules['PROVIDER_EPG'])
                    ERROR = True

                else:
                    pass

            except:
                ERROR = True
                OUTPUT_LOG.append(
                    {'Errors': 'EPG "' + rules['PROVIDER_EPG'] + '" does not conform to the naming standard'})

        DISPLAY_SET = set(DISPLAY_LIST)
        for contracts in DISPLAY_SET:
            OUTPUT_LOG.append({'Errors': 'EPG "' + contracts + '" does not conform to the naming standard'})
        DISPLAY_LIST = []

    if not ERROR:
        OUTPUT_LOG.append({'NotificationsSuccess': 'EPG formatting validated successfully'})

    OUTPUT_LOG.append({'Headers': 'Validating Contract and EPG locality.'})

    for rules in RULE_LIST:

        if rules['CONSUMER_EPG'] != 'BLANK':
            if rules['CONSUMER_EPG'].split('-')[0].upper() != rules['NAME'].split('-')[0].upper():
                DISPLAY_LIST.append(rules['CONSUMER_EPG'])
                ERROR = True

        if rules['PROVIDER_EPG'] != 'BLANK':
            if rules['PROVIDER_EPG'].split('-')[0].upper() != rules['NAME'].split('-')[0].upper():
                DISPLAY_LIST.append(rules['PROVIDER_EPG'])
                ERROR = True

        else:
            pass

    DISPLAY_SET = set(DISPLAY_LIST)
    for contracts in DISPLAY_SET:
        OUTPUT_LOG.append({'Errors': contracts + ' and Contract named for different Tenants.'})
    DISPLAY_LIST = []

    if not ERROR:
        OUTPUT_LOG.append({'NotificationsSuccess': 'Contract and EPG locality validated successfully'})

    OUTPUT_LOG.append({'Headers': 'Validating Services in Workbook.'})

    for rules in RULE_LIST:
        for services in rules['SERVICE']:
            PROTOCOLS = ['TCP', 'UDP']
            try:
                value = int(services.split('-')[1])
            except:
                OUTPUT_LOG.append(
                    {'Errors': 'Service port not correct format for ' + services + ' on line ' + str(rules['LINE'])})
            if services.split('-')[0] not in PROTOCOLS:
                OUTPUT_LOG.append({'Errors': 'TCP or UDP not specified for service ' + services})
                ERROR = True

            elif int(services.split('-')[1]) == 0:
                OUTPUT_LOG.append({'Errors': 'Error Port out of range ' + services})
                ERROR = True

            elif int(services.split('-')[1]) not in range(1, 65536):
                OUTPUT_LOG.append({'Errors': 'Error Port out of range ' + services})
                ERROR = True

            elif len(services.split('-')) == 3:
                if int(services.split('-')[2]) not in range(1, 65536):
                    OUTPUT_LOG.append({'Errors': 'Error Port out of range ' + services})
                    ERROR = True

            else:
                pass

    if not ERROR:
        OUTPUT_LOG.append({'NotificationsSuccess': 'Services validated successfully'})

        OUTPUT_LOG.append({'Headers': 'Connecting to APIC'})
        APIC_COOKIE = APIC_LOGIN(BASE_URL, username, password)

        if APIC_COOKIE:
            OUTPUT_LOG.append({'Notifications': 'Successfully generated authentication cookie'})

            # Check if Contracts Exist
            OUTPUT_LOG.append({'Headers': 'Checking if Contracts exist'})
            for rules in RULE_LIST:
                CONTRACT_NAME = rules['NAME']
                CONTRACT_SEARCH_RESPONSE = CONTRACT_SEARCH(BASE_URL, APIC_COOKIE, CONTRACT_NAME, HEADERS)
                if int(CONTRACT_SEARCH_RESPONSE['totalCount']) == 1:
                    pass
                else:
                    CONTRACT_LIST.append(CONTRACT_NAME)

            CONTRACT_SET = set(CONTRACT_LIST)
            for contracts in CONTRACT_SET:
                OUTPUT_LOG.append({'Notifications': 'Contract ' + contracts + ' will be created'})

            # Check if Filters Exist
            OUTPUT_LOG.append({'Headers': 'Checking if Filters exist'})
            for rules in RULE_LIST:
                for services in rules['SERVICE']:
                    FILTER_NAME = services
                    FILTER_SEARCH_RESPONSE = FILTER_SEARCH(BASE_URL, APIC_COOKIE, FILTER_NAME, HEADERS)
                    if int(FILTER_SEARCH_RESPONSE['totalCount']) == 1:
                        pass
                    else:
                        FILTER_LIST.append(FILTER_NAME)

            FILTER_SET = set(FILTER_LIST)
            for filters in FILTER_SET:
                OUTPUT_LOG.append({'Notifications': 'Filter ' + filters + ' will be created'})

            # Check if Filters are applied to contracts
            OUTPUT_LOG.append({'Headers': 'Checking if Filters are applied to contracts'})
            for rules in RULE_LIST:
                # Use the contract search to locate subject
                CONTRACT_NAME = rules['NAME']
                CONTRACT_SEARCH_RESPONSE = CONTRACT_SEARCH(BASE_URL, APIC_COOKIE, CONTRACT_NAME, HEADERS)
                # Validate that a contract can be located
                if int(CONTRACT_SEARCH_RESPONSE['totalCount']) == 1:
                    TENANT = CONTRACT_SEARCH_RESPONSE['imdata'][0]['vzBrCP']['attributes']['dn'].split('/')[1][3:]
                    CONTRACT_SUBJECT = \
                        CONTRACT_SEARCH_RESPONSE['imdata'][0]['vzBrCP']['attributes']['dn'].split('/')[2][4:].split(
                            '_')[0] + '_SBJ'
                    SUBJECT_SEARCH_RESPONSE = SUBJECT_SEARCH(BASE_URL, APIC_COOKIE, CONTRACT_NAME, TENANT,
                                                             CONTRACT_SUBJECT,
                                                             HEADERS)

                    # Add all filters for a subject to a list to be used for comparison.
                    SUBJECT_FILTERS = []
                    for filters in SUBJECT_SEARCH_RESPONSE['imdata']:
                        SUBJECT_FILTERS.append(filters['vzRsSubjFiltAtt']['attributes']['tnVzFilterName'])

                    # compare list of filters to those already in subject

                    FILTER_COMPARE = LIST_COMPARE(rules['SERVICE'], SUBJECT_FILTERS)
                    if len(FILTER_COMPARE[0]) == 0:
                        pass

                    elif len(FILTER_COMPARE[0]) > 0:
                        OUTPUT_LOG.append(
                            {'Notifications': 'The below filters will be added to contract ' + CONTRACT_NAME})
                        OUTPUT_LOG.append({'Notifications': FILTER_COMPARE[0]})

                    else:
                        ERROR = True
                else:
                    pass

            # Check if L3Outs Exist
            OUTPUT_LOG.append({'Headers': 'Checking if L3Outs exist'})
            for rules in RULE_LIST:
                if rules['CONSUMER_L3OUT'] == 'INTERNAL':
                    pass
                else:
                    L3OUT_NAME = rules['CONSUMER_L3OUT']
                    L3OUT_SEARCH_RESPONSE = CONTRACT_L3OUT_SEARCH(BASE_URL, APIC_COOKIE, L3OUT_NAME, HEADERS)
                    if int(L3OUT_SEARCH_RESPONSE['totalCount']) == 1:
                        if rules['CONSUMER_L3OUT'] == L3OUT_SEARCH_RESPONSE['imdata'][0]['l3extOut']['attributes'][
                            'name']:
                            pass
                        else:
                            ERROR = True

                    else:
                        DISPLAY_LIST.append(L3OUT_NAME)
                        ERROR = True

            for rules in RULE_LIST:
                if rules['PROVIDER_L3OUT'] == 'INTERNAL':
                    pass
                else:
                    L3OUT_NAME = rules['PROVIDER_L3OUT']
                    L3OUT_SEARCH_RESPONSE = CONTRACT_L3OUT_SEARCH(BASE_URL, APIC_COOKIE, L3OUT_NAME, HEADERS)
                    if int(L3OUT_SEARCH_RESPONSE['totalCount']) == 1:
                        if rules['PROVIDER_L3OUT'] == L3OUT_SEARCH_RESPONSE['imdata'][0]['l3extOut']['attributes'][
                            'name']:
                            pass
                        else:
                            ERROR = True

                    else:
                        DISPLAY_LIST.append(L3OUT_NAME)
                        ERROR = True

            DISPLAY_SET = set(DISPLAY_LIST)
            for l3out in DISPLAY_SET:
                OUTPUT_LOG.append({'Errors': 'L3Out: ' + l3out + ' Does not exist, please check naming.'})
            DISPLAY_LIST = []

            if not ERROR:
                # Check if External EPGs Exist
                OUTPUT_LOG.append({'Headers': 'Checking if External EPGs are created'})
                for rules in RULE_LIST:
                    if rules['CONSUMER_L3OUT'] == 'INTERNAL':
                        pass
                    else:
                        if rules['CONSUMER_EPG'] != 'BLANK':
                            CONSUMER_L3OUT = rules['CONSUMER_L3OUT']
                            CONSUMER_EPG = rules['CONSUMER_EPG']
                            EXTERNAL_EPG_SEARCH_RESPONSE = CONTRACT_EXTERNAL_EPG_SEARCH(BASE_URL, APIC_COOKIE,
                                                                                        CONSUMER_L3OUT, CONSUMER_EPG,
                                                                                        HEADERS)
                            if int(EXTERNAL_EPG_SEARCH_RESPONSE['totalCount']) == 1:
                                EPG_NAME = \
                                EXTERNAL_EPG_SEARCH_RESPONSE['imdata'][0]['l3extInstP']['attributes']['dn'].split('/')[
                                    3][6:]
                                L3OUT_NAME = \
                                EXTERNAL_EPG_SEARCH_RESPONSE['imdata'][0]['l3extInstP']['attributes']['dn'].split('/')[
                                    2][4:]
                                EXTERNAL_EPG_SEARCH_RESPONSE['imdata'][0]['l3extInstP']['attributes']['dn'].split('/')[
                                    1][3:]
                                if L3OUT_NAME == rules['CONSUMER_L3OUT']:
                                    pass
                                else:
                                    OUTPUT_LOG.append({
                                                          'Errors': 'EPG and L3OUT missmatch with ' + L3OUT_NAME + ' and ' + EPG_NAME + ' dont match value: ' +
                                                                    rules['CONSUMER_L3OUT']})

                            else:
                                DISPLAY_LIST.append(CONSUMER_EPG)
                                ERROR = True

                for rules in RULE_LIST:
                    PROVIDER_EPG = rules['PROVIDER_EPG']
                    if rules['PROVIDER_L3OUT'] == 'INTERNAL':
                        pass
                    else:
                        if rules['PROVIDER_EPG'] != 'BLANK':
                            PROVIDER_EPG = rules['PROVIDER_EPG']
                            PROVIDER_L3OUT = rules['PROVIDER_L3OUT']
                            EXTERNAL_EPG_SEARCH_RESPONSE = CONTRACT_EXTERNAL_EPG_SEARCH(BASE_URL, APIC_COOKIE,
                                                                                        PROVIDER_L3OUT, PROVIDER_EPG,
                                                                                        HEADERS)
                            if int(EXTERNAL_EPG_SEARCH_RESPONSE['totalCount']) == 1:
                                EPG_NAME = \
                                EXTERNAL_EPG_SEARCH_RESPONSE['imdata'][0]['l3extInstP']['attributes']['dn'].split('/')[
                                    3][6:]
                                L3OUT_NAME = \
                                EXTERNAL_EPG_SEARCH_RESPONSE['imdata'][0]['l3extInstP']['attributes']['dn'].split('/')[
                                    2][4:]
                                EXTERNAL_EPG_SEARCH_RESPONSE['imdata'][0]['l3extInstP']['attributes']['dn'].split('/')[
                                    1][3:]
                                if L3OUT_NAME == rules['PROVIDER_L3OUT']:
                                    pass
                                else:
                                    OUTPUT_LOG.append({
                                                          'Errors': 'EPG and L3OUT missmatch with ' + L3OUT_NAME + ' and ' + EPG_NAME + ' dont match value: ' +
                                                                    rules['PROVIDER_L3OUT']})

                            else:
                                DISPLAY_LIST.append(PROVIDER_EPG)
                                ERROR = True

                DISPLAY_SET = set(DISPLAY_LIST)
                for epgs in DISPLAY_SET:
                    OUTPUT_LOG.append({'Errors': 'EPG "' + epgs + '" needs creating.'})

            # Check if Internal EPGs Exist
            OUTPUT_LOG.append({'Headers': 'Checking if Internal EPGs are created'})
            EPG_LIST = []
            # Search to validate internal EPG's are created
            for rules in RULE_LIST:
                if rules['CONSUMER_L3OUT'] == 'INTERNAL':
                    if rules['CONSUMER_EPG'] != 'BLANK':
                        CONSUMER_EPG = rules['CONSUMER_EPG']
                        # Search for consumer EPG
                        INTERNAL_EPG_SEARCH_RESPONSE = INTERNAL_EPG_SEARCH(BASE_URL, APIC_COOKIE, CONSUMER_EPG, HEADERS)
                        if int(INTERNAL_EPG_SEARCH_RESPONSE['totalCount']) == 1:
                            pass
                        else:
                            DISPLAY_LIST.append(CONSUMER_EPG)
                            ERROR = True
                else:
                    pass

            for rules in RULE_LIST:
                if rules['PROVIDER_L3OUT'] == 'INTERNAL':
                    if rules['PROVIDER_EPG'] != 'BLANK':
                        PROVIDER_EPG = rules['PROVIDER_EPG']
                        # Search for Provider EPG
                        INTERNAL_EPG_SEARCH_RESPONSE = INTERNAL_EPG_SEARCH(BASE_URL, APIC_COOKIE, PROVIDER_EPG, HEADERS)
                        if int(INTERNAL_EPG_SEARCH_RESPONSE['totalCount']) == 1:
                            pass
                        else:
                            DISPLAY_LIST.append(PROVIDER_EPG)
                            ERROR = True
                else:
                    pass

            DISPLAY_SET = set(DISPLAY_LIST)
            for epgs in DISPLAY_SET:
                OUTPUT_LOG.append({'Errors': 'EPG "' + epgs + '" needs creating.'})
            DISPLAY_LIST = []


        else:
            OUTPUT_LOG.append({'Errors': 'Unable to connect to APIC.'})

    if not ERROR:
        OUTPUT_LOG.append({'ValidationSuccess': 'APIC Configuration validated successfully'})

    return OUTPUT_LOG


@shared_task
def CONTRACT_DEPLOYMENT(RULE_LIST, location, url_dict, username, password):
    BASE_URL = url_dict[location]

    CONTRACT_LIST = []
    FILTER_LIST = []
    DISPLAY_LIST = []
    OUTPUT_LOG = []
    HEADERS = {'content-type': 'application/json'}
    # --------------------------------------------------------------------------#
    # Begin Configuration
    # --------------------------------------------------------------------------#
    OUTPUT_LOG.append({'Headers': 'Starting contract provisioning.'})
    OUTPUT_LOG.append({'Headers': 'Connecting to APIC'})
    APIC_COOKIE = APIC_LOGIN(BASE_URL, username, password)
    if APIC_COOKIE:
        OUTPUT_LOG.append({'Notifications': 'Successfully generated authentication cookie'})

        OUTPUT_LOG.append({'Headers': 'Creating Filters.'})
        # Create Filters in Filter SET
        for rules in RULE_LIST:
            for services in rules['SERVICE']:
                FILTER_NAME = services
                FILTER_SEARCH_RESPONSE = FILTER_SEARCH(BASE_URL, APIC_COOKIE, FILTER_NAME, HEADERS)
                if int(FILTER_SEARCH_RESPONSE['totalCount']) == 1:
                    pass
                else:
                    FILTER_LIST.append(FILTER_NAME)

        FILTER_SET = set(FILTER_LIST)
        if len(FILTER_SET) > 0:
            OUTPUT_LOG = FILTER_CREATE(FILTER_SET, BASE_URL, APIC_COOKIE, HEADERS, OUTPUT_LOG)
        else:
            pass

        OUTPUT_LOG.append({'Headers': 'Creating Contract & Subjects.'})
        # Create Contracts and Subjects in Contract SET
        for rules in RULE_LIST:
            CONTRACT_NAME = rules['NAME']
            CONTRACT_SEARCH_RESPONSE = CONTRACT_SEARCH(BASE_URL, APIC_COOKIE, CONTRACT_NAME, HEADERS)
            if int(CONTRACT_SEARCH_RESPONSE['totalCount']) == 1:
                pass
            else:
                CONTRACT_LIST.append(CONTRACT_NAME)

        CONTRACT_SET = set(CONTRACT_LIST)
        if len(CONTRACT_SET) > 0:
            OUTPUT_LOG = CONTRACT_CREATE(CONTRACT_SET, BASE_URL, APIC_COOKIE, HEADERS, OUTPUT_LOG)
        else:
            pass

        # Add filters to subjects
        for rules in RULE_LIST:
            # Use the contract search to locate subject
            CONTRACT_NAME = rules['NAME']
            CONTRACT_SEARCH_RESPONSE = CONTRACT_SEARCH(BASE_URL, APIC_COOKIE, CONTRACT_NAME, HEADERS)
            # Validate that a contract can be located
            if int(CONTRACT_SEARCH_RESPONSE['totalCount']) == 1:
                TENANT = CONTRACT_SEARCH_RESPONSE['imdata'][0]['vzBrCP']['attributes']['dn'].split('/')[1][3:]
                CONTRACT_SUBJECT = \
                    CONTRACT_SEARCH_RESPONSE['imdata'][0]['vzBrCP']['attributes']['dn'].split('/')[2][4:].split('_')[
                        0] + '_SBJ'
                SUBJECT_SEARCH_RESPONSE = SUBJECT_SEARCH(BASE_URL, APIC_COOKIE, CONTRACT_NAME, TENANT, CONTRACT_SUBJECT,
                                                         HEADERS)

                # Add all filters for a subject to a list to be used for comparison.
                SUBJECT_FILTERS = []
                for filters in SUBJECT_SEARCH_RESPONSE['imdata']:
                    SUBJECT_FILTERS.append(filters['vzRsSubjFiltAtt']['attributes']['tnVzFilterName'])

                # compare list of filters to those already in subject
                NEW_FILTERS = rules['SERVICE']
                FILTER_COMPARE = LIST_COMPARE(NEW_FILTERS, SUBJECT_FILTERS)
                if len(FILTER_COMPARE[0]) == 0:
                    pass

                elif len(FILTER_COMPARE[0]) > 0:
                    for FILTERS in FILTER_COMPARE[0]:
                        OUTPUT_LOG = FILTER_ATTACH(BASE_URL, APIC_COOKIE, CONTRACT_NAME, CONTRACT_SUBJECT, FILTERS,
                                                   HEADERS,
                                                   OUTPUT_LOG)

                else:
                    OUTPUT_LOG.append({'Errors': 'Error adding Filters to subject!'})

                NEW_FILTERS = ''

            else:
                pass

        OUTPUT_LOG.append({'Headers': 'Consuming & Providing Contracts.'})

        for rules in RULE_LIST:
            LINE = rules['LINE']
            CONTRACT_NAME = rules['NAME']
            # Consuming contract
            EPG_NAME = rules['CONSUMER_EPG']
            OUTPUT_LOG.append({'Headers2': 'Deploying Contracts for Line: ' + str(LINE)})

            if EPG_NAME != 'BLANK':
                if rules['CONSUMER_L3OUT'] == 'INTERNAL':
                    INTERNL_EPG_CONTRACT_CONSUME(BASE_URL, EPG_NAME, CONTRACT_NAME, APIC_COOKIE, HEADERS, OUTPUT_LOG)
                else:
                    L3OUT_NAME = rules['CONSUMER_L3OUT']
                    EXTERNAL_EPG_CONTRACT_CONSUME(L3OUT_NAME, EPG_NAME, CONTRACT_NAME, BASE_URL, APIC_COOKIE, HEADERS,
                                                  OUTPUT_LOG)

            # Providing contract
            EPG_NAME = rules['PROVIDER_EPG']
            if EPG_NAME != 'BLANK':
                if rules['PROVIDER_L3OUT'] == 'INTERNAL':
                    INTERNL_EPG_CONTRACT_PROVIDE(BASE_URL, EPG_NAME, CONTRACT_NAME, APIC_COOKIE, HEADERS, OUTPUT_LOG)
                else:
                    L3OUT_NAME = rules['PROVIDER_L3OUT']
                    EXTERNAL_EPG_CONTRACT_PROVIDE(L3OUT_NAME, EPG_NAME, CONTRACT_NAME, BASE_URL, APIC_COOKIE, HEADERS,
                                                  OUTPUT_LOG)

    else:
        OUTPUT_LOG.append({'Errors': 'Unable to connect to APIC.'})

    return OUTPUT_LOG
