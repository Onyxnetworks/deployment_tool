# Base Functions
import openpyxl, ipaddress, netaddr
from netaddr import IPNetwork, IPAddress
# Celery Functions
from celery import shared_task

# Custom Functions
from f5_deployment.scripts.vs_deployment import *
from f5_deployment.scripts.generic_search import *
from f5_deployment.scripts.enable_disable import *

def vs_deployment_excel_open_workbook(file):
    wb = openpyxl.load_workbook(file, data_only=True)
    py_ws = wb.get_sheet_by_name('Tab_Python')
    vs_dict = {}
    vs_dict['version'] = str(py_ws['CD1'].value)
    vs_dict['vs'] = {}
    vs_dict['node_list'] = []
    vs_dict['node_priority'] = []
    for col in py_ws.iter_cols(min_row=2, max_row=2):
        for cell in col:
            vs_dict['vs'][cell.coordinate] = cell.value

    for col in py_ws.iter_cols(min_row=2, max_row=2, min_col=28, max_col=47):
        for cell in col:
            cell = str(cell.value)
            if not cell == 'None':
                vs_dict['node_list'].append(cell)
            
    for col in py_ws.iter_cols(min_row=2, max_row=2, min_col=49, max_col=68):
        for cell in col:
            cell = str(cell.value)
            if not cell == 'None':        
                vs_dict['node_priority'].append(cell)        

    return vs_dict
    
@shared_task
def vs_deployment_validation(vs_dict, location, url_dict, username, password):
    
    output_log = []
    error = False
    form_version = vs_dict['version']
    script_version = '1.4'
    output_log.append({'Headers': 'Checking Script compatibility.'})
    if form_version != script_version:
        error = True
        output_log.append(
            {'Errors': 'Script version missmatch. Excel version: {0} | Script version: {1}'.format(form_version, script_version)})
        return output_log, vs_dict
    else:
        output_log.append({'NotificationsSuccess': 'Form version validated successfully.'})


    output_log.append({'Headers': 'Identifying F5 Device group.'})  
    # Get Virtual Server Name
    vs_name = vs_dict['vs']['A2']

    if location == 'UKDC1':
        device_group = vs_name.rsplit('-', 10)[0] + '-DGA'
        base_url = url_dict[location][device_group]
        output_log.append({'Notifications': 'Device group ' + device_group + ' DGA selected.'})
        
    elif location == 'UKDC2':
        device_group = vs_name.rsplit('-', 10)[0] + '-DGB'
        base_url = url_dict[location][device_group]
        output_log.append({'Notifications': 'Device group ' + device_group + ' DGB selected.'})
    
    elif location == 'LAB':
        device_group = 'LAB'
        base_url = url_dict[location][device_group]
        output_log.append({'Notifications': 'Device group ' + device_group + ' selected.'})
    else:
        error = True
        output_log.append({'Errors': 'Unable to identify F5 Device group, please check index/baseline.py configuration.'})
        return output_log, vs_dict

    if not error:   
        output_log.append({'Headers': 'Creating connection to BigIP.'})
        bigip_connection = create_connection_bigip(base_url, username, password, output_log)
        output_log = bigip_connection[0]
        error = bigip_connection[1]

    if not error:
        output_log.append({'Headers': 'Validating Partition.'})
        try:
            bigip_url_base = bigip_connection[2]
            bigip = bigip_connection[3]
            partition = vs_dict['vs']['E2']
            validate_partition_result = validate_partition(partition, bigip_url_base, bigip, output_log)
            output_log = validate_partition_result[0]
            error = validate_partition_result[1]

        except:
            error = True
            output_log.append(
                {'Errors': 'Unable to identify partition.'})
            return output_log, vs_dict

    if not error:
        output_log.append({'Headers': 'Checking Sync status of F5.'})   
        #check_sync_result  = check_sync(bigip_url_base, bigip, output_log)
        #output_log = check_sync_result[0]
        #error = check_sync_result[1]
     
    if not error:
        output_log.append({'Headers': 'Checking for HTTP Profiles.'})   
        check_httpprofile_result = check_httpprofile(vs_dict, partition, bigip_url_base, bigip, output_log)
        
        output_log = check_httpprofile_result[0]
        error = check_httpprofile_result[1]
       
    if not error:
        output_log.append({'NotificationsSuccess': 'HTTP Profiles validated successfully.'})
        ssl = vs_dict['vs']['D2']
        if ssl in ['https', 'HTTPS']:
                output_log.append({'Headers': 'Checking SSL Profiles.'})
                check_ssl_profile_result = check_ssl_profile(vs_dict, partition, bigip_url_base, bigip, output_log)
                output_log = check_ssl_profile_result[0]
                error = check_ssl_profile_result[1]
                if not error:
                    output_log.append({'Headers': 'Checking certificates.'})
                    check_cert_result = check_cert(vs_dict, partition, bigip_url_base, bigip, output_log)
                    output_log = check_cert_result[0]
                    error = check_cert_result[1]
    
    if not error:
        output_log.append({'Headers': 'Checking HTTP(S) Monitors.'})
        check_http_mon_result = check_http_mon(vs_dict, partition, bigip_url_base, bigip, output_log)
        output_log = check_http_mon_result[0]
        error = check_http_mon_result[1]
    
    if not error:
        output_log.append({'NotificationsSuccess': 'Monitor configuration validated successfully.'})
        output_log.append({'Headers': 'Checking SNAT Pools.'})
        snat_pool_present = compare_snat_on_ltm_excel(vs_dict, partition, bigip_url_base, bigip, output_log)
        output_log = snat_pool_present[0]
        error = snat_pool_present[1]
        snat_pool_present = snat_pool_present[2]
        vs_dict['snat_pool_present'] = snat_pool_present
        
    if not error:
        output_log.append({'NotificationsSuccess': 'SNAT configuration validated successfully.'})
        output_log.append({'Headers': 'Checking Nodes.'})
        node_list_result = compare_ltm_nodes(vs_dict, partition, bigip_url_base, bigip, output_log)
        output_log = node_list_result[0]
        error = node_list_result[1]

    if not error:
        node_list = node_list_result[2]
        node_list_pool = node_list_result[3]
        vs_dict['node_list'] = node_list
        vs_dict['node_list_pool'] = node_list_pool

    if not error:
        output_log.append({'NotificationsSuccess': 'Node configuration validated successfully.'})
        output_log.append({'Headers': 'Checking Pools.'})
        compare_pool_results = compare_pool(vs_dict, partition, bigip_url_base, bigip, output_log)
        output_log = compare_pool_results[0]
        error = compare_pool_results[1]

    if not error:
        output_log.append({'NotificationsSuccess': 'Pool configuration validated successfully.'})
        output_log.append({'Headers': 'Checking Virtual Server.'})
        compare_vs_results = compare_vs(vs_dict, partition, bigip_url_base, bigip, output_log)
        output_log = compare_vs_results[0]
        error = compare_vs_results[1]

    if not error:
        output_log.append({'NotificationsSuccess': 'Virtual Server configuration validated successfully.'})
        output_log.append({'ValidationSuccess': 'LTM Configuration validated successfully'})






    return output_log, vs_dict


@shared_task
def virtual_server_deployment(vs_dict, location, url_dict, routeAdvertisement, username, password):

    error = False
    output_log = []
    output_log.append({'Headers': 'Starting Virtual Server deployment.'})

    vs_name = vs_dict['vs']['A2']
    partition = vs_dict['vs']['E2']

    if location == 'UKDC1':
        device_group = vs_name.rsplit('-', 10)[0] + '-DGA'
        base_url = url_dict[location][device_group]

    elif location == 'UKDC2':
        device_group = vs_name.rsplit('-', 10)[0] + '-DGB'
        base_url = url_dict[location][device_group]

    elif location == 'LAB':
        device_group = 'LAB'
        base_url = url_dict[location][device_group]

    else:
        error = True
        output_log.append(
            {'Errors': 'Unable to identify F5 Device group, please check index/baseline.py configuration.'})

    if not error:
        output_log.append({'Headers': 'Creating connection to BigIP.'})
        bigip_connection = create_connection_bigip(base_url, username, password, output_log)
        output_log = bigip_connection[0]
        error = bigip_connection[1]

    if not error:
        bigip_url_base = bigip_connection[2]
        bigip = bigip_connection[3]

        if (vs_dict['vs']['Y2']) and (vs_dict['vs']['L2']):
            create_vs_ssl_profiles_result = create_vs_ssl_profiles(vs_dict, partition, bigip_url_base, bigip, output_log)
            output_log = create_vs_ssl_profiles_result[0]
            error = create_vs_ssl_profiles_result[1]

        # Skip Monitor deployment if TCP is the type.
        if not vs_dict['vs']['Q2'] == 'tcp':
            if not error:
                create_pool_monitor_result = create_pool_monitor(vs_dict, partition, bigip_url_base, bigip, output_log)
                output_log = create_pool_monitor_result[0]
                error = create_pool_monitor_result[1]

        if not error:
            create_vs_profiles_http_result = create_vs_profiles_http(vs_dict, partition, bigip_url_base, bigip, output_log)
            output_log = create_vs_profiles_http_result[0]
            error = create_vs_profiles_http_result[1]

        if not error:
            if vs_dict['snat_pool_present'] == 0:
                create_snat_result = create_snat(vs_dict, partition, bigip_url_base, bigip, output_log)
                output_log = create_snat_result[0]
                error = create_snat_result[1]

        if not error:
            if vs_dict['node_list']:
                node_list = vs_dict['node_list']
                create_nodes_result = create_nodes(node_list, partition, bigip_url_base, bigip, output_log)
                output_log = create_nodes_result[0]
                error = create_nodes_result[1]

        if not error:
            create_pool_result = create_pool(vs_dict, partition, bigip_url_base, bigip, output_log)
            output_log = create_pool_result[0]
            error = create_pool_result[1]

        if not error:

            create_vs_result = create_vs(vs_dict, partition, bigip_url_base, bigip, output_log)
            output_log = create_vs_result[0]
            error = create_vs_result[1]

        if not error:
            if routeAdvertisement:
                create_vs_result = create_advertise_vip(vs_dict, partition, bigip_url_base, bigip, output_log)
                output_log = create_vs_result[0]
                error = create_vs_result[1]
            else:
                output_log.append(
                    {'Notifications': 'Selected Virtual Server IP to not be advertised.'})

        if not error:
            output_log.append({'ValidationSuccess': 'Virtual Server deployed successfully.'})


    return output_log


@shared_task
def f5_generic_search(base_urls, role, request_type, search_string, username, password):
    # Build URL List to search.
    url_list = []
    base_urls = list(base_urls['F5'].values())
    for urls in base_urls:
        for url in urls.values():
            url_list.append(url)
    
    if request_type == 'Virtual Server Name':
        # Get virtual server name
        results = []
        virtual_server_dashboard_result = virtual_server_dashboard(url_list, request_type, search_string, username, password)
        if isinstance(virtual_server_dashboard_result, list):
            results = {'search': {'search_string': search_string, 'role': role, 'request_type': request_type}, 'data': virtual_server_dashboard_result}

            return results


    if request_type == 'Virtual Server IP':
        # Get virtual server IP
        results = []
        virtual_server_dashboard_result = virtual_server_dashboard(url_list, request_type, search_string, username, password)
        if isinstance(virtual_server_dashboard_result, list):
            for vs in virtual_server_dashboard_result:
                try:
                    network = ipaddress.IPv4Network(search_string)
                    results = {'search': {'search_string': search_string, 'request_type': request_type},
                               'data': virtual_server_dashboard_result}

                except ValueError:
                    pass

        return results

    if request_type == 'Pool':
        # Get Pool name
        results = []
        virtual_server_dashboard_result = virtual_server_dashboard(url_list, request_type, search_string, username, password)
        if isinstance(virtual_server_dashboard_result, list):
            results = {'search': {'search_string': search_string, 'request_type': request_type}, 'data': virtual_server_dashboard_result}

            return results


@shared_task
def f5_disable_enable(base_urls, request_type, action, f5_selected_items, username, password):
    # Different json posts for various F5 states
    node_forced_offline_json = {"state": "user-down", "session": "user-disabled"}
    node_disabled_json = {"state": "user-up", "session": "user-disabled"}
    node_enabled_json = {"state": "user-up", "session": "user-enabled"}
    vs_disabled_json = {"disabled": True}
    vs_enabled_json = {"enabled": True}
    # Loop through items and disable them on thee respective F5.
    for selflink in f5_selected_items:
        # get URL got base login (hostname without)
        login_url = selflink.split('/mgmt/')[0]
        bigip_login_response = bigip_login(login_url, username, password)
        auth_token = bigip_login_response['token']['token']

        if action == 'disable' and request_type == 'node':
            disable_response = node_disable_enable_force(selflink, auth_token, node_disabled_json)

        if action == 'enable' and request_type == 'node':
            enable_response = node_disable_enable_force(selflink, auth_token, node_enabled_json)

        if action == 'disable' and request_type == 'vs':
            disable_response = vs_disable_enable_force(selflink, auth_token, vs_disabled_json)

        if action == 'enable' and request_type == 'vs':
            enable_response = vs_disable_enable_force(selflink, auth_token, vs_enabled_json)


@shared_task
def certificate_checker_task(base_urls, role, request_type, search_string, username, password):
    # Build URL List to search.
    url_list = []
    base_urls = list(base_urls['F5'].values())
    for urls in base_urls:
        for url in urls.values():
            url_list.append(url)


    results = []
    certificate_checker_result = certificate_checker(url_list, request_type, search_string, username, password)

    if isinstance(certificate_checker_result, list):
        results = {'search': {'search_string': search_string, 'role': role, 'request_type': request_type},
                   'data': certificate_checker_result}

        return results





